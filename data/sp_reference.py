"""
S&P (Land-BRZ25) reference price lookup.

Source: LAND-BRZ25 Brazil Farmland Market Analysis Ed.117, Jan-Mar 2025.
Prices are in R$/ha as of March 2025.

Lookup hierarchy:
  1. Exact (UF, municipio, subgrupo) match              → match_level = "municipio"
  2. Region average via T17 annex (region_id, subgrupo) → match_level = "regiao"
  3. State-level average (UF, subgrupo)                 → match_level = "estado"
  4. None (no data available)

For each level, low/mid/high are derived as:
  - low  = average of all CAPACIDADE="Baixa" prices at that level
  - high = average of all CAPACIDADE="Alta" prices at that level
  - mid  = average of ALL prices (all capacidades) at that level
  (Média tier fills in when Alta or Baixa is absent)

T17 annex maps every Brazilian municipality to a Region ID so that
when a city is not directly in "Base de dados consolidada" we can
still find a meaningful regional benchmark rather than falling all
the way back to a state-wide average.
"""

from __future__ import annotations

import unicodedata
import os
from collections import defaultdict
from functools import lru_cache
from typing import Optional

# ---------------------------------------------------------------------------
# Land type → S&P subgrupo mapping
# ---------------------------------------------------------------------------
_LANDTYPE_TO_SUBGRUPO: dict[str, list[str]] = {
    "soja":          ["Grãos", "Produção Diversificada"],
    "grãos":         ["Grãos", "Produção Diversificada"],
    "lavoura":       ["Grãos", "Produção Diversificada"],
    "milho":         ["Grãos", "Produção Diversificada"],
    "cana":          ["Cana"],
    "café":          ["Café"],
    "cafe":          ["Café"],
    "arroz":         ["Arroz"],
    "fruticultura":  ["Fruticultura"],
    "pastagem":      ["Pastagem"],
    "pecuária":      ["Pastagem"],
    "pecuaria":      ["Pastagem"],
    "mata":          ["Cerrado", "Floresta Amazonica", "Mata Atlântica"],
    "floresta":      ["Florestas Plantadas", "Floresta Amazonica"],
    "silvicultura":  ["Florestas Plantadas"],
    "misto":         ["Produção Diversificada", "Grãos", "Pastagem"],
    # fallback
    "_default":      ["Grãos", "Produção Diversificada", "Pastagem"],
}


def _normalize(s: str) -> str:
    """Strip accents and lowercase for fuzzy matching."""
    return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode().lower().strip()


# ---------------------------------------------------------------------------
# File location
# ---------------------------------------------------------------------------
_XLSX_FILENAME = "LAND-BRZ25-Brazil_Farmland_Market_Analysis_Ed_117_Jan-Mar_2025_English-Portuguese_tables_last_udpated_on_4-15-25_v2.xlsx"


def _find_xlsx() -> Optional[str]:
    candidates = [
        os.path.join(os.path.dirname(__file__), "reference", _XLSX_FILENAME),
        os.path.join(os.path.dirname(__file__), "..", "..", "project 2", _XLSX_FILENAME),
        os.path.expanduser(f"~/Desktop/project 2/{_XLSX_FILENAME}"),
        os.path.expanduser(f"~/Desktop/{_XLSX_FILENAME}"),
    ]
    for p in candidates:
        p = os.path.normpath(p)
        if os.path.exists(p):
            return p
    return None


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

@lru_cache(maxsize=1)
def _load_sp_data() -> dict:
    """
    Load S&P xlsx and return indexed lookup dicts.

    Returns dict with keys:
      'mun_sub'   : {(uf, norm_mun, subgrupo)  -> {cap -> [prices]}}
      'reg_sub'   : {(region_id, subgrupo)      -> {cap -> [prices]}}
      'state_sub' : {(uf, subgrupo)             -> {cap -> [prices]}}
      't17'       : {(norm_mun, uf)             -> (region_id, region_name)}
      'rows'      : list of raw row dicts for the reference table
    """
    path = _find_xlsx()
    if path is None:
        return {"mun_sub": {}, "reg_sub": {}, "state_sub": {}, "t17": {}, "rows": []}

    try:
        import openpyxl
    except ImportError:
        return {"mun_sub": {}, "reg_sub": {}, "state_sub": {}, "t17": {}, "rows": []}

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # ── Base de dados consolidada ────────────────────────────────────────────
    # Header row index 3; data starts at index 4.
    # Columns: ID_CLASSN(0) ID_REG(1) REGIAO(2) UF(3) ID_MUN(4) MUNICIPIO(5)
    #          GRUPO(6) SUBGRUPO(7) USO(8) CAPACIDADE(9) COMPLEMENTO(10)
    #          VALOR_UNI(11) UNI(12) BIOMA(13) STATUS(14) [blank](15)
    #          price_2001(16) price_dec24(17) price_mar25(18)
    ws = wb["Base de dados consolidada"]
    raw_rows = list(ws.iter_rows(values_only=True))
    wb_rows = raw_rows[4:]  # skip title + header rows

    mun_sub:   dict = defaultdict(lambda: defaultdict(list))
    reg_sub:   dict = defaultdict(lambda: defaultdict(list))
    state_sub: dict = defaultdict(lambda: defaultdict(list))
    rows_out = []

    for r in wb_rows:
        if r[14] != 1:
            continue
        region_id = r[1]
        regiao    = r[2]
        uf        = r[3]
        mun       = r[5]
        subgrupo  = r[7]
        cap       = r[9]
        price     = r[18]   # Mar 2025

        if not all([uf, mun, subgrupo, cap, price]):
            continue
        if not isinstance(price, (int, float)):
            continue

        uf        = str(uf).strip().upper()
        mun       = str(mun).strip()
        sub       = str(subgrupo).strip()
        cap       = str(cap).strip()
        region_id = int(region_id) if isinstance(region_id, (int, float)) else None

        mun_key = (uf, _normalize(mun), sub)
        mun_sub[mun_key][cap].append(price)

        if region_id is not None:
            reg_sub[(region_id, sub)][cap].append(price)

        state_key = (uf, sub)
        state_sub[state_key][cap].append(price)

        rows_out.append({
            "region_id": region_id,
            "regiao":    str(regiao).strip() if regiao else "",
            "uf":        uf,
            "municipio": mun,
            "subgrupo":  sub,
            "capacidade": cap,
            "price_mar25": price,
        })

    # ── T17 — municipality → region mapping ──────────────────────────────────
    # Header row at index 5; data starts at index 6.
    # Cols: [1]=State/UF  [2]=ID Region  [3]=Region  [4]=Cód.Mun.IBGE  [5]=Municipality
    wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws17 = wb2["T17"]
    t17_rows = list(ws17.iter_rows(values_only=True))
    wb2.close()

    t17: dict = {}
    for r in t17_rows[6:]:   # skip rows 0-5 (blanks + header)
        uf17  = r[1]
        rid   = r[2]
        rname = r[3]
        mun17 = r[5]
        if not all([uf17, rid, rname, mun17]):
            continue
        if rid == "ND":
            continue
        uf17  = str(uf17).strip().upper()
        mun17 = str(mun17).strip()
        rid   = int(rid) if isinstance(rid, (int, float)) else None
        if rid is None:
            continue
        key = (_normalize(mun17), uf17)
        t17[key] = (rid, str(rname).strip())

    return {
        "mun_sub":   dict(mun_sub),
        "reg_sub":   dict(reg_sub),
        "state_sub": dict(state_sub),
        "t17":       t17,
        "rows":      rows_out,
    }


# ---------------------------------------------------------------------------
# Aggregate capacidade tiers into (low, mid, high)
# ---------------------------------------------------------------------------

def _agg(cap_dict: dict) -> tuple[Optional[float], float, Optional[float]]:
    """
    Given {cap -> [prices]}, return (baixa_avg, mid_avg, alta_avg).
    mid = average of ALL prices regardless of tier.
    baixa / alta fall back to Média tier when absent.
    """
    def avg(lst):
        return round(sum(lst) / len(lst)) if lst else None

    baixa = avg(cap_dict.get("Baixa", []))
    media = avg(cap_dict.get("Média", []))
    alta  = avg(cap_dict.get("Alta",  []))

    all_prices = [p for lst in cap_dict.values() for p in lst]
    mid = avg(all_prices)

    return (baixa or media, mid, alta or media)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def get_sp_reference(
    uf: str,
    city: str,
    land_type: str,
) -> Optional[tuple[float, float, float, str]]:
    """
    Return (low, mid, high, match_level) R$/ha from S&P data.

    match_level:
      "municipio" — exact municipality match in Base de dados consolidada
      "regiao"    — municipality found in T17; figures are the regional average
      "estado"    — state-wide average fallback
    """
    data = _load_sp_data()
    if not data["mun_sub"]:
        return None

    uf        = str(uf).strip().upper()
    city_norm = _normalize(str(city))
    lt_lower  = land_type.lower().strip()

    subgrupos = _LANDTYPE_TO_SUBGRUPO.get(lt_lower, _LANDTYPE_TO_SUBGRUPO["_default"])

    # 1. Exact municipality match
    for sub in subgrupos:
        key = (uf, city_norm, sub)
        if key in data["mun_sub"]:
            baixa, mid, alta = _agg(data["mun_sub"][key])
            if mid:
                return (baixa, mid, alta, "municipio")

    # 2. Region average via T17
    t17_key = (city_norm, uf)
    region_id = None
    region_name = ""
    if t17_key in data["t17"]:
        region_id, region_name = data["t17"][t17_key]

    if region_id is not None:
        for sub in subgrupos:
            key = (region_id, sub)
            if key in data["reg_sub"]:
                baixa, mid, alta = _agg(data["reg_sub"][key])
                if mid:
                    return (baixa, mid, alta, "regiao")

    # 3. State-level fallback
    for sub in subgrupos:
        key = (uf, sub)
        if key in data["state_sub"]:
            baixa, mid, alta = _agg(data["state_sub"][key])
            if mid:
                return (baixa, mid, alta, "estado")

    return None


# ---------------------------------------------------------------------------
# Reference table for dashboard tab
# ---------------------------------------------------------------------------

_MATCH_LEVEL_LABELS = {
    "municipio": "Município",
    "regiao":    "Região",
    "estado":    "Média Estadual",
}

_STATE_AVG_LABEL    = "— Média estadual —"
_REGION_AVG_PREFIX  = "— Média região: "


def sp_reference_table() -> "pd.DataFrame":
    """
    Return the full S&P reference dataset as a tidy DataFrame.

    One row per (uf, municipio, subgrupo) — capacidade tiers collapsed.
    Plus one "— Média região: X —" row per (region_id, subgrupo).
    Plus one "— Média estadual —" row per (uf, subgrupo).

    row_type: "municipio" | "regiao" | "estado"
    """
    import pandas as pd

    data = _load_sp_data()
    if not data["mun_sub"]:
        return pd.DataFrame()

    STATE_NAMES = {
        "AC": "Acre", "AL": "Alagoas", "AM": "Amazonas", "AP": "Amapá",
        "BA": "Bahia", "CE": "Ceará", "ES": "Espírito Santo", "GO": "Goiás",
        "MA": "Maranhão", "MG": "Minas Gerais", "MS": "Mato Grosso do Sul",
        "MT": "Mato Grosso", "PA": "Pará", "PB": "Paraíba", "PE": "Pernambuco",
        "PI": "Piauí", "PR": "Paraná", "RJ": "Rio de Janeiro", "RN": "Rio Grande do Norte",
        "RO": "Rondônia", "RR": "Roraima", "RS": "Rio Grande do Sul",
        "SC": "Santa Catarina", "SE": "Sergipe", "SP": "São Paulo", "TO": "Tocantins",
    }

    # ── Municipality rows ────────────────────────────────────────────────────
    from collections import defaultdict
    grouped: dict = defaultdict(lambda: {"region_id": None, "regiao": "", "uf": "", "caps": defaultdict(list)})
    for row in data["rows"]:
        k = (row["uf"], row["municipio"], row["subgrupo"])
        grouped[k]["region_id"] = row["region_id"]
        grouped[k]["regiao"]    = row["regiao"]
        grouped[k]["uf"]        = row["uf"]
        grouped[k]["caps"][row["capacidade"]].append(row["price_mar25"])

    out = []
    for (uf, mun, sub), v in grouped.items():
        baixa, mid, alta = _agg(v["caps"])
        out.append({
            "region_id":   v["region_id"],
            "regiao":      v["regiao"],
            "uf":          uf,
            "state_name":  STATE_NAMES.get(uf, uf),
            "municipio":   mun,
            "subgrupo":    sub,
            "price_baixa": baixa,
            "price_mid":   mid,
            "price_alta":  alta,
            "row_type":    "municipio",
        })

    # ── Region-average rows ──────────────────────────────────────────────────
    # Need region_id → (uf set, region_name) for labelling
    reg_meta: dict = {}
    for row in data["rows"]:
        rid = row["region_id"]
        if rid and rid not in reg_meta:
            reg_meta[rid] = {"regiao": row["regiao"], "ufs": set()}
        if rid:
            reg_meta[rid]["ufs"].add(row["uf"])

    for (rid, sub), cap_dict in data["reg_sub"].items():
        baixa, mid, alta = _agg(cap_dict)
        meta = reg_meta.get(rid, {})
        rname = meta.get("regiao", str(rid))
        ufs = sorted(meta.get("ufs", []))
        uf_label = ufs[0] if len(ufs) == 1 else "/".join(ufs)
        out.append({
            "region_id":   rid,
            "regiao":      rname,
            "uf":          uf_label,
            "state_name":  STATE_NAMES.get(uf_label, uf_label),
            "municipio":   f"{_REGION_AVG_PREFIX}{rname} —",
            "subgrupo":    sub,
            "price_baixa": baixa,
            "price_mid":   mid,
            "price_alta":  alta,
            "row_type":    "regiao",
        })

    # ── State-average rows ───────────────────────────────────────────────────
    for (uf, sub), cap_dict in data["state_sub"].items():
        baixa, mid, alta = _agg(cap_dict)
        out.append({
            "region_id":   None,
            "regiao":      "",
            "uf":          uf,
            "state_name":  STATE_NAMES.get(uf, uf),
            "municipio":   _STATE_AVG_LABEL,
            "subgrupo":    sub,
            "price_baixa": baixa,
            "price_mid":   mid,
            "price_alta":  alta,
            "row_type":    "estado",
        })

    df = pd.DataFrame(out)
    df["_sort_type"] = df["row_type"].map({"municipio": 0, "regiao": 1, "estado": 2})
    df = (
        df.sort_values(["uf", "subgrupo", "_sort_type", "municipio"])
          .drop(columns="_sort_type")
          .reset_index(drop=True)
    )
    return df
